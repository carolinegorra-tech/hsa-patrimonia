"""
GENERIC PATRIMON.IA — LISTA DE ATIVOS (EXCEL) BUILDER
======================================================
build_excel(data: dict, output_path: str)

Fills the HSA "Lista de Ativos" Excel deliverable for ANY client.

Design: instead of inserting/deleting rows in the template (fragile —
formula refs drift and you get #VALUE!/#REF! in the totals), this builder
COMPUTES THE FULL LAYOUT FIRST, then writes every row from scratch into a
fresh sheet, copying the template's cell styles by row-type. Every SUM /
section-total formula is generated from the final, known row numbers, so
there is never a stale or self-referential range.

The template (Lista_Ativos_TEMPLATE_BLANK_1.xlsx) is used only as the
STYLE SOURCE — we read its fonts/fills/number-formats, never its layout.
"""
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

TEMPLATE = '/mnt/project/Lista_Ativos_TEMPLATE_BLANK_1.xlsx'

# ── Category mapping: template block → keywords matching client group names
# IMPORTANT: order matters (first match wins) and keywords must be specific
# enough not to collide — e.g. "ações" as a bare substring also matches
# "particip-ações" and "aplic-ações", so the Ações block uses distinctive
# stock terms instead, and the more specific blocks are tested first.
BR_BLOCKS = [
    ("Imóveis",                     ['imóve', 'imove']),
    ("Bens Móveis / Obras de Arte", ['móve', 'move', 'arte', 'veícul', 'veicul']),
    ("Participações Societárias",   ['ociet', 'participa', 'quota', 'cota']),
    ("Fundos de Investimento",      ['fundo']),
    ("Investimentos Renda Fixa",    ['renda fixa', 'renda-fixa', 'renda fix',
                                     'cdb', 'lci', 'lca', 'tesouro', 'aplica']),
    ("Contas Bancárias",            ['conta', 'banc', 'corrente', 'poupanç', 'poupanc']),
    ("Ações",                       ['ação', 'ações', 'bolsa', 'petr', 'vale',
                                     'itub', 'bbas', 'papéis', 'papeis', 'b3']),
    ("Créditos e Direitos",         ['crédit', 'credit', 'direito']),
]
OFF_BLOCKS = [
    ("Participações Societárias no Exterior",     ['ociet', 'participa', 'capital', 'empresa']),
    ("Depósitos e Contas Bancárias no Exterior",  ['conta', 'banc', 'depósit', 'deposit', 'corrente']),
    ("Seguros e Previdência Privada no Exterior", ['seguro', 'previd']),
]

def _match(group_name, blocks, fallback_idx):
    g = group_name.lower()
    for bn, kws in blocks:
        if any(k in g for k in kws):
            return bn
    return blocks[fallback_idx][0]


# ════════════════════════════════════════════════════════════════════════════
# STYLE LIBRARY — read once from the template, keyed by row-type
# ════════════════════════════════════════════════════════════════════════════
def _grab_styles(template_path):
    wb = load_workbook(template_path)
    ws = wb.active
    def snap(coord):
        c = ws[coord]
        return {
            'font': copy(c.font),
            'fill': copy(c.fill),
            'border': copy(c.border),
            'alignment': copy(c.alignment),
            'number_format': c.number_format,
        }
    styles = {
        'title':       snap('C2'),
        'subtitle':    snap('C3'),
        'hdr_pais':    snap('D6'),
        'hdr_dirpf':   snap('E6'),
        'hdr_dcbe':    snap('F6'),
        'hdr_coment':  snap('G6'),
        'section':     snap('B8'),    # BRASIL / OFFSHORE bar
        'subsection':  snap('B9'),    # "Bens e Direitos"
        'block':       snap('B10'),   # "Imóveis"
        'item_num':    snap('B11'),
        'item_desc':   snap('C11'),
        'item_pais':   snap('D11'),
        'item_dirpf':  snap('E11'),
        'item_dcbe':   snap('F11'),
        'item_coment': snap('G11'),
        'btot_label':  snap('B13'),   # block total label
        'btot_dirpf':  snap('E13'),
        'btot_dcbe':   snap('F48'),
        'stot_label':  snap('B42'),   # section total label
        'stot_dirpf':  snap('E42'),
        'grand_label': snap('B59'),
        'grand_value': snap('E59'),
        'footnote':    snap('B68'),
        'col_widths':  {k: v.width for k, v in ws.column_dimensions.items()},
    }

    # ── derived USD-formatted variants ──────────────────────────────────────
    # The template has no USD-formatted section-total or grand-value cells
    # (F42/F59 are 'General'). Build them by cloning the BRL style and swapping
    # the number_format to the USD pattern (taken from btot_dcbe = F48).
    USD_FMT = styles['btot_dcbe']['number_format']  # '"US$ "#,##0.00'
    def _usd_variant(base):
        return {
            'font':          copy(base['font']),
            'fill':          copy(base['fill']),
            'border':        copy(base['border']),
            'alignment':     copy(base['alignment']),
            'number_format': USD_FMT,
        }
    styles['stot_dcbe']       = _usd_variant(styles['stot_dirpf'])
    styles['grand_value_usd'] = _usd_variant(styles['grand_value'])
    return styles


def _apply(cell, style):
    cell.font = copy(style['font'])
    cell.fill = copy(style['fill'])
    cell.border = copy(style['border'])
    cell.alignment = copy(style['alignment'])
    cell.number_format = style['number_format']


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def build_excel(data, output_path):
    S = _grab_styles(TEMPLATE)

    CLIENT = data.get('client', '—')
    YEAR   = data.get('year', '')
    groups = data.get('groups', [])
    debts  = data.get('debts', []) or []

    # ── bucket client items into template blocks ────────────────────────────
    br_items  = {b[0]: [] for b in BR_BLOCKS}
    off_items = {b[0]: [] for b in OFF_BLOCKS}
    for g in groups:
        juris = g.get('jurisdiction', 'Brasil')
        name  = g.get('name', '')
        if juris == 'Brasil':
            tgt = br_items[_match(name, BR_BLOCKS, 7)]
        else:
            tgt = off_items[_match(name, OFF_BLOCKS, 0)]
        for it in g.get('items', []):
            tgt.append({
                'desc':     it.get('desc', ''),
                'loc':      it.get('loc', '') or ('Brasil' if juris == 'Brasil' else 'Exterior'),
                'dirpf':    it.get('dirpf'),
                'dcbe':     it.get('dcbe'),
                'comments': it.get('comments', ''),
            })

    # ── fresh workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f'Lista de Ativos 31.12.{YEAR}'

    for col, w in S['col_widths'].items():
        if w:
            ws.column_dimensions[col].width = w

    item_counter = [0]

    def write_row(rr, *, b=None, c=None, d=None, e=None, f=None, g=None, styles=None):
        vals = {'B': b, 'C': c, 'D': d, 'E': e, 'F': f, 'G': g}
        for col, val in vals.items():
            cell = ws[f'{col}{rr}']
            if val is not None:
                cell.value = val
            if styles and col in styles:
                _apply(cell, S[styles[col]])

    # ── HEADER ──────────────────────────────────────────────────────────────
    write_row(2, c=CLIENT, styles={'C': 'title'})
    write_row(3, c=f'LISTA DE ATIVOS (DIRPF e DCBE {YEAR} - AC {YEAR})',
              styles={'C': 'subtitle'})
    write_row(6,
              d='País',
              e=f'Valor DIRPF\nano-calendário {YEAR}',
              f=f'Valor DCBE\nano-calendário {YEAR}',
              g='Comentários HSA',
              styles={'D':'hdr_pais','E':'hdr_dirpf','F':'hdr_dcbe','G':'hdr_coment'})

    r = 8
    # Per-column block-total rows. Brasil writes only E; Offshore writes both E and F.
    block_total_rows_br_e  = []
    block_total_rows_off_e = []
    block_total_rows_off_f = []

    def emit_block(block_name, items, write_dirpf, write_dcbe,
                   store_e, store_f):
        """
        write_dirpf: emit DIRPF (col E) item values + block total in E
        write_dcbe : emit DCBE  (col F) item values + block total in F
        Brasil   blocks pass (True, False, store_e, None)
        Offshore blocks pass (True, True,  store_e, store_f)
        """
        nonlocal r
        write_row(r, b=block_name, styles={k:'block' for k in 'BCDEFG'})
        r += 1
        first_item = r
        wrote = 0
        for it in items:
            item_counter[0] += 1
            write_row(r,
                      b=item_counter[0],
                      c=it['desc'],
                      d=it['loc'],
                      e=(it['dirpf'] if write_dirpf else None),
                      f=(it['dcbe']  if write_dcbe  else None),
                      g=it['comments'],
                      styles={'B':'item_num','C':'item_desc','D':'item_pais',
                              'E':'item_dirpf','F':'item_dcbe','G':'item_coment'})
            r += 1
            wrote += 1
        if wrote == 0:
            write_row(r, d=('Brasil' if write_dirpf and not write_dcbe else 'Exterior'),
                      styles={'B':'item_num','C':'item_desc','D':'item_pais',
                              'E':'item_dirpf','F':'item_dcbe','G':'item_coment'})
            r += 1
            wrote = 1
        first, last = first_item, r - 1
        # Block-total row — apply value style to whichever columns receive a sum
        write_row(r,
                  b=f'Total {block_name}:',
                  styles={'B':'btot_label','C':'btot_label','D':'btot_label',
                          'E':('btot_dirpf' if write_dirpf else 'btot_label'),
                          'F':('btot_dcbe'  if write_dcbe  else 'btot_label'),
                          'G':'btot_label'})
        if write_dirpf:
            ws[f'E{r}'] = f'=SUM(E{first}:E{last})'
            if store_e is not None:
                store_e.append(r)
        if write_dcbe:
            ws[f'F{r}'] = f'=SUM(F{first}:F{last})'
            if store_f is not None:
                store_f.append(r)
        r += 1

    # ── BRASIL ──────────────────────────────────────────────────────────────
    write_row(r, b='BRASIL', styles={k:'section' for k in 'BCDEFG'})
    ws.merge_cells(f'B{r}:G{r}'); r += 1
    write_row(r, b='Bens e Direitos', styles={k:'subsection' for k in 'BCDEFG'})
    ws.merge_cells(f'B{r}:G{r}'); r += 1
    for bn, _ in BR_BLOCKS:
        # Brasil: only DIRPF column has values + totals
        emit_block(bn, br_items[bn],
                   write_dirpf=True, write_dcbe=False,
                   store_e=block_total_rows_br_e, store_f=None)

    write_row(r, b='Total Ativos Brasil:',
              styles={'B':'stot_label','C':'stot_label','D':'stot_label',
                      'E':'stot_dirpf','F':'stot_label','G':'stot_label'})
    ws.merge_cells(f'B{r}:D{r}')
    ws[f'E{r}'] = '=' + '+'.join(f'E{x}' for x in block_total_rows_br_e)
    brasil_total_row = r
    r += 2

    # ── OFFSHORE ────────────────────────────────────────────────────────────
    write_row(r, b='OFFSHORE', styles={k:'section' for k in 'BCDEFG'})
    ws.merge_cells(f'B{r}:G{r}'); r += 1
    for bn, _ in OFF_BLOCKS:
        # Offshore: items carry BOTH DIRPF (BRL, col E) and DCBE (USD, col F).
        # PTAX is informational only — values are stored as declared, never converted here.
        emit_block(bn, off_items[bn],
                   write_dirpf=True, write_dcbe=True,
                   store_e=block_total_rows_off_e, store_f=block_total_rows_off_f)

    write_row(r, b='Total Ativos Offshore:',
              styles={'B':'stot_label','C':'stot_label','D':'stot_label',
                      'E':'stot_dirpf','F':'stot_dcbe','G':'stot_label'})
    ws.merge_cells(f'B{r}:D{r}')
    if block_total_rows_off_e:
        ws[f'E{r}'] = '=' + '+'.join(f'E{x}' for x in block_total_rows_off_e)
    if block_total_rows_off_f:
        ws[f'F{r}'] = '=' + '+'.join(f'F{x}' for x in block_total_rows_off_f)
    offshore_total_row = r
    r += 2

    # ── TOTAL ATIVOS ────────────────────────────────────────────────────────
    # Col E (BRL): combined DIRPF — Brasil + offshore items' declared DIRPF value.
    # Col F (USD): DCBE total — offshore only.
    write_row(r, b='TOTAL ATIVOS',
              styles={'B':'grand_label','C':'grand_label','D':'grand_label',
                      'E':'grand_value','F':'grand_value_usd','G':'grand_label'})
    ws.merge_cells(f'B{r}:D{r}')
    ws[f'E{r}'] = f'=E{brasil_total_row}+E{offshore_total_row}'
    ws[f'F{r}'] = f'=F{offshore_total_row}'
    r += 2

    # ── DÍVIDAS ─────────────────────────────────────────────────────────────
    write_row(r, b='Dívidas e Ônus Reais', styles={k:'subsection' for k in 'BCDEFG'})
    ws.merge_cells(f'B{r}:G{r}'); r += 1
    write_row(r, b='Dívidas', styles={k:'block' for k in 'BCDEFG'})
    r += 1
    debt_first = r
    if debts:
        for d in debts:
            item_counter[0] += 1
            write_row(r,
                      b=item_counter[0],
                      c=d.get('desc', ''),
                      d='Brasil',
                      e=d.get('value'),
                      styles={'B':'item_num','C':'item_desc','D':'item_pais',
                              'E':'item_dirpf','F':'item_dcbe','G':'item_coment'})
            r += 1
    else:
        write_row(r, d='Brasil',
                  styles={'B':'item_num','C':'item_desc','D':'item_pais',
                          'E':'item_dirpf','F':'item_dcbe','G':'item_coment'})
        r += 1
    debt_last = r - 1
    write_row(r, b='Total Dívidas:',
              styles={'B':'btot_label','C':'btot_label','D':'btot_label',
                      'E':'btot_dirpf','F':'btot_label','G':'btot_label'})
    ws[f'E{r}'] = f'=SUM(E{debt_first}:E{debt_last})'
    debt_total_row = r
    r += 1
    write_row(r, b='TOTAL DÍVIDAS E ÔNUS REAIS:',
              styles={'B':'grand_label','C':'grand_label','D':'grand_label',
                      'E':'grand_value','F':'grand_label','G':'grand_label'})
    ws.merge_cells(f'B{r}:D{r}')
    ws[f'E{r}'] = f'=E{debt_total_row}'
    r += 2

    # ── FOOTNOTE ────────────────────────────────────────────────────────────
    write_row(r, b=(f'Fonte: DIRPF {YEAR} (AC {YEAR}) — Receita Federal do Brasil  |  '
                    f'DCBE Anual {YEAR} (Data-base 31/12/{YEAR}) — Banco Central do Brasil'),
              styles={k:'footnote' for k in 'BCDEFG'})
    ws.merge_cells(f'B{r}:G{r}')

    # row heights
    for rr in range(1, r + 1):
        if ws.row_dimensions[rr].height is None:
            ws.row_dimensions[rr].height = 15.0
    ws.row_dimensions[2].height = 23.25
    ws.row_dimensions[6].height = 28.5

    wb.save(output_path)
    return True


# ════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import json
    BFM = json.loads(r'''{"client":"BEATRIZ FERREIRA MONTEIRO","cpf":"987.654.321-00","year":2024,"spouse":{"name":"RODRIGO ANDRADE MONTEIRO"},"dependents":[],"groups":[{"name":"Bens Imóveis","jurisdiction":"Brasil","items":[{"id":1,"desc":"Total Grupo 01 — Bens Imóveis (Apartamento Itaim Bibi SP + Casa de praia Juquehy SP)","loc":"São Paulo, SP / Juquehy, SP","dirpf":5150000,"dcbe":null,"comments":"Grupo 01 DIRPF — total oficial do resumo"}]},{"name":"Bens Móveis / Veículos / Arte","jurisdiction":"Brasil","items":[{"id":2,"desc":"Total Grupo 02 — Bens Móveis / Veículos / Arte (BMW X5 2023 + Obras de arte 3 peças)","loc":"São Paulo, SP","dirpf":1310000,"dcbe":null,"comments":"Grupo 02 DIRPF — total oficial do resumo"}]},{"name":"Participações Societárias","jurisdiction":"Brasil","items":[{"id":3,"desc":"Total Grupo 03 — Participações Societárias (PETR4 + Monteiro Empreendimentos 60% + BFM Participacoes 100%)","loc":"Brasil","dirpf":29250000,"dcbe":null,"comments":"Grupo 03 DIRPF — total oficial do resumo"}]},{"name":"Aplicações e Investimentos","jurisdiction":"Brasil","items":[{"id":4,"desc":"Total Grupo 04 — Aplicações e Investimentos (CDB Itaú BBA + Fundo Bradesco FIC FIM + LCI XP)","loc":"Brasil","dirpf":5920000,"dcbe":null,"comments":"Grupo 04 DIRPF — total oficial do resumo"}]},{"name":"Contas Bancárias","jurisdiction":"Brasil","items":[{"id":5,"desc":"Total Grupo 06 — Contas Bancárias (Itaú Unibanco + Nubank)","loc":"Brasil","dirpf":99300,"dcbe":null,"comments":"Grupo 06 DIRPF — total oficial do resumo"}]},{"name":"Participações no Capital de Empresas no Exterior","jurisdiction":"Offshore","items":[{"id":6,"desc":"BFM INTERNATIONAL LTD. — 100% participação — método Patrimônio Líquido","loc":"Ilhas Cayman","dirpf":null,"dcbe":3250000,"comments":"DCBE Ficha A"},{"id":7,"desc":"MONTEIRO HOLDINGS LLC — 50% participação — método Patrimônio Líquido","loc":"Delaware, EUA","dirpf":null,"dcbe":1880000,"comments":"DCBE Ficha A"}]},{"name":"Contas Bancárias no Exterior","jurisdiction":"Offshore","items":[{"id":8,"desc":"Bank of America N.A. — Conta Corrente USD","loc":"EUA","dirpf":null,"dcbe":285400,"comments":"DCBE Ficha B"},{"id":9,"desc":"Bank of America N.A. — Conta Poupança USD","loc":"EUA","dirpf":null,"dcbe":112000,"comments":"DCBE Ficha B"},{"id":10,"desc":"Banco Santander España — Conta Corrente EUR","loc":"Espanha","dirpf":null,"dcbe":202585,"comments":"DCBE Ficha B — EUR convertido PTAX 1,0389"}]}],"debts":[{"id":1,"desc":"Financiamento imobiliário — Banco Bradesco S.A. CNPJ 60.746.948/0001-12 — Cód 16","value":1095000}]}''')
    ok = build_excel(BFM, '/home/claude/BFM_2024_LISTA_ATIVOS.xlsx')
    print("✓ Build OK" if ok else "✗ FAILED")
